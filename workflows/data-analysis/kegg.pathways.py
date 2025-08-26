#!/usr/bin/env python3
"""
Create a KEGG pathway spreadsheet with the exact
indentation/layout requested by the user.

 • Column A: headers only (1., 1.0, 1.1, …)
 • Column B: pathway ID + flags (M/R/N/T)
 • Column C: pathway name
 • Blank row inserted between each L1 and its first L2,
   so '1.0 Global and overview maps' sits two rows below
   '1. Metabolism' just like the example.
"""
import re
import requests
import openpyxl
from openpyxl.utils import get_column_letter

URL = "https://www.genome.jp/kegg/pathway.html"
html = requests.get(URL, timeout=60).text.splitlines()

# ------------------------------------------------------------------
# Regular-expression helpers
# ------------------------------------------------------------------
rx_l1 = re.compile(r"^\s*####\s+(\d+\.\s+[A-Z].+)$")          # 1. Metabolism, 2. Genetic...
rx_l2 = re.compile(r"^\s*(\d+\.\d+\s+[A-Z].+)$")              # 1.0 Global..., 1.1 Carb...
rx_code = re.compile(r"^\s*(\d{5}(?:\s+[MNRT](?:\s+[MNRT])*)?)\s*$")  # 01100 M, 01200 M R...

# Keep track of section state so we can insert blank row after each L1
waiting_for_first_l2 = False

wb = openpyxl.Workbook()
ws = wb.active
row = 1   # Excel rows start at 1

def write(cell_col, value):
    """Helper to write to current row, given column index (1=A,2=B,3=C)."""
    if value is None: 
        return
    ws[f"{get_column_letter(cell_col)}{row}"] = value

# ------------------------------------------------------------------
# Two-pass scan: KEGG’s HTML is simple — every pathway code line is
# followed immediately by an anchor line with the name.
# ------------------------------------------------------------------
for i, line in enumerate(html):
    # 1) Top-level header (#### 1. Metabolism …)
    m = rx_l1.match(line)
    if m:
        # Write header
        if row > 1:   # not the very first header → add a blank line before it
            row += 1
        write(1, m.group(1).strip())  # Column A
        row += 1                      # Advance to next (blank) row
        waiting_for_first_l2 = True
        continue

    # 2) Second-level header (1.0 Global …, 1.1 Carbohydrate …, 1.2 Energy …, etc.)
    m = rx_l2.match(line)
    if m:
        if waiting_for_first_l2:
            # leave this row blank so header sits two lines below L1
            row += 1
            waiting_for_first_l2 = False
        write(1, m.group(1).strip())  # Column A
        row += 1
        continue

    # 3) Pathway code line (five-digit number ± flags)
    m = rx_code.match(line)
    if m:
        code_with_flags = m.group(1)
        # Next line (i+1) contains the anchor with the pathway title → fetch it
        if i + 1 < len(html):
            name_line = html[i + 1]
            # strip the Unicode bracket/anchor junk, leaving just the title text
            name = re.sub(r"^.*?】", "", name_line).strip()
        else:
            name = ""
        write(2, code_with_flags)     # Column B
        write(3, name)                # Column C
        row += 1

# ------------------------------------------------------------------
# Finish and save
# ------------------------------------------------------------------
wb.save("kegg_pathways_formatted.xlsx")
print(f"Done!  Wrote {row-1} lines ➜ kegg_pathways_formatted.xlsx")
